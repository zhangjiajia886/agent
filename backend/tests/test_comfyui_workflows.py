#!/usr/bin/env python
"""
远程 ComfyUI Workflow 可用性测试
绕过 Agent，直接调用 ComfyUI API，测试每个 workflow 是否可用
并记录耗时，写回 DB test_time 字段

用法: conda run -n ttsapp python tests/test_comfyui_workflows.py
输出: /tmp/comfyui_workflow_test_report.md
"""
import asyncio
import copy
import json
import os
import random
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import httpx

# ── 配置 ──────────────────────────────────────────────
COMFYUI_URL = "https://u982127-7772b8fbe6d9.bjb1.seetacloud.com:8443"
COMFYUI_TIMEOUT = 900  # 最长等待秒数（LTX2 等大模型首次加载需 >600s）
POLL_INTERVAL = 3.0

# 测试素材（用已有的 agent_outputs 图片）
BACKEND_DIR = Path(__file__).resolve().parent.parent
UPLOADS_DIR = BACKEND_DIR / "uploads"
AGENT_OUTPUTS_DIR = UPLOADS_DIR / "agent_outputs"
WORKFLOWS_ROOT = BACKEND_DIR / "app" / "core" / "comic_agent" / "workflows"

# DB 配置
DB_HOST = "localhost"
DB_PORT = 3306
DB_USER = "ttsapp"
DB_PASS = "ttsapp123"
DB_NAME = "ttsapp"

# 输出
REPORT_PATH = "/tmp/comfyui_workflow_test_report.md"
EXCEL_PATH = "/tmp/comfyui_workflow_test_report.xlsx"
LOG_PATH = "/tmp/comfyui_workflow_test.log"


# ── 数据结构 ──────────────────────────────────────────────
@dataclass
class WorkflowTestCase:
    db_id: int
    db_name: str
    display_name: str
    category: str  # t2i, edit, face, i2v, t2v, upscale, audio
    style_tag: Optional[str]
    needs_image: bool = False
    needs_audio: bool = False
    workflow_json: Optional[dict] = None
    # 结果
    status: str = "pending"  # pending, pass, fail, skip
    error: str = ""
    duration_s: float = 0.0
    output_type: str = ""  # image, video, audio
    output_size: int = 0


def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# ── ComfyUI 客户端（独立实现，不依赖后端代码） ────────────
class SimpleComfyUIClient:
    def __init__(self, base_url: str):
        self.base_url = base_url.rstrip("/")
        self.client_id = f"test-{random.randint(1000,9999)}"

    async def health_check(self) -> tuple[bool, str]:
        try:
            async with httpx.AsyncClient(timeout=15, verify=False) as c:
                r = await c.get(f"{self.base_url}/system_stats")
                if r.status_code == 200:
                    data = r.json()
                    # 获取 GPU 信息
                    devices = data.get("devices", [])
                    gpu_info = ", ".join(d.get("name", "?") for d in devices) if devices else "unknown"
                    return True, gpu_info
                return False, f"HTTP {r.status_code}"
        except Exception as e:
            return False, str(e)

    async def upload_image(self, image_bytes: bytes, filename: str = "test_input.png") -> str:
        async with httpx.AsyncClient(timeout=30, verify=False) as c:
            r = await c.post(
                f"{self.base_url}/upload/image",
                files={"image": (filename, image_bytes, "image/png")},
            )
            r.raise_for_status()
            return r.json()["name"]

    async def submit_workflow(self, workflow: dict) -> str:
        clean = {k: v for k, v in workflow.items() if not k.startswith("__")}
        payload = {"prompt": clean, "client_id": self.client_id}
        async with httpx.AsyncClient(timeout=30, verify=False) as c:
            r = await c.post(f"{self.base_url}/prompt", json=payload)
            if r.status_code != 200:
                try:
                    body = r.json()
                    err_msg = body.get("error", {}).get("message", "")
                    node_errors = body.get("node_errors", {})
                    details = "; ".join(
                        f"{nid}({nd.get('class_type','')}): {[e.get('details','') for e in nd.get('errors',[])]}"
                        for nid, nd in node_errors.items()
                    )
                    raise RuntimeError(f"ComfyUI {r.status_code}: {err_msg} | {details}")
                except RuntimeError:
                    raise
                except Exception:
                    raise RuntimeError(f"ComfyUI HTTP {r.status_code}: {r.text[:200]}")
            result = r.json()
            if "error" in result:
                raise RuntimeError(f"提交失败: {result['error']}")
            if result.get("node_errors"):
                errs = result["node_errors"]
                details = "; ".join(
                    f"{nid}({nd.get('class_type','')}): {[e.get('details','') for e in nd.get('errors',[])]}"
                    for nid, nd in errs.items()
                )
                raise RuntimeError(f"节点错误: {details}")
            return result["prompt_id"]

    async def wait_result(self, prompt_id: str) -> dict:
        deadline = time.time() + COMFYUI_TIMEOUT
        async with httpx.AsyncClient(timeout=15, verify=False) as c:
            while time.time() < deadline:
                r = await c.get(f"{self.base_url}/history/{prompt_id}")
                data = r.json()
                if prompt_id in data:
                    status_info = data[prompt_id].get("status", {})
                    if status_info.get("status_str") == "error":
                        msgs = status_info.get("messages", [])
                        # 提取 execution_error 中的关键信息
                        err_detail = "unknown execution error"
                        for msg in msgs:
                            if isinstance(msg, list) and len(msg) >= 2 and msg[0] == "execution_error":
                                d = msg[1]
                                err_detail = f"{d.get('node_id','?')}({d.get('node_type','?')}): {d.get('exception_type','?')}: {d.get('exception_message','?')[:200]}"
                                break
                        raise RuntimeError(f"执行失败: {err_detail}")
                    return data[prompt_id].get("outputs", {})
                await asyncio.sleep(POLL_INTERVAL)
        raise TimeoutError(f"超时 ({COMFYUI_TIMEOUT}s)")

    async def download_output(self, filename: str, subfolder: str = "", file_type: str = "output") -> bytes:
        async with httpx.AsyncClient(timeout=60, verify=False) as c:
            r = await c.get(
                f"{self.base_url}/view",
                params={"filename": filename, "subfolder": subfolder, "type": file_type},
            )
            r.raise_for_status()
            return r.content

    async def run_workflow(self, workflow: dict) -> tuple[str, bytes]:
        """返回 (output_type, bytes)"""
        prompt_id = await self.submit_workflow(workflow)
        log(f"    提交成功 prompt_id={prompt_id[:8]}... 等待结果...")
        outputs = await self.wait_result(prompt_id)
        for node_output in outputs.values():
            for media_key, media_type in [("images", "image"), ("videos", "video"), ("gifs", "video"), ("audio", "audio")]:
                if media_key in node_output:
                    item = node_output[media_key][0]
                    data = await self.download_output(
                        item["filename"],
                        subfolder=item.get("subfolder", ""),
                        file_type=item.get("type", "output"),
                    )
                    return media_type, data
        raise RuntimeError("无输出（无 images/videos/audio）")


# ── Workflow JSON 加载与参数注入 ────────────────────────
_WF_CACHE: dict[str, Path] | None = None


def _get_wf_cache() -> dict[str, Path]:
    global _WF_CACHE
    if _WF_CACHE is None:
        _WF_CACHE = {}
        for p in WORKFLOWS_ROOT.rglob("*.json"):
            rel = p.relative_to(WORKFLOWS_ROOT)
            s = str(rel.with_suffix(""))
            name = s.replace("/", "__").replace(" ", "_").replace(".", "_")[:100]
            _WF_CACHE[name] = p
    return _WF_CACHE


def load_workflow_json(wf_name: str) -> dict:
    """按 DB name 加载 workflow JSON（使用 name→path 缓存）"""
    # 1. 直接文件名（旧格式）
    simple = WORKFLOWS_ROOT / f"{wf_name}.json"
    if simple.exists():
        return json.loads(simple.read_text(encoding="utf-8"))
    # 2. 精确匹配缓存
    cache = _get_wf_cache()
    if wf_name in cache:
        return json.loads(cache[wf_name].read_text(encoding="utf-8"))
    raise FileNotFoundError(f"Workflow JSON not found: {wf_name}")


def inject_test_params(
    workflow: dict,
    positive_prompt: str = "a beautiful girl standing in a garden, masterpiece, best quality",
    negative_prompt: str = "ugly, deformed, blurry, bad anatomy, watermark, text, low quality, nsfw",
    seed: int = 42,
    width: int = 512,
    height: int = 512,
    source_image: Optional[str] = None,
    instruction: Optional[str] = None,
) -> dict:
    """注入测试参数（简化版 inject_params）"""
    wf = copy.deepcopy(workflow)
    node_items = [(k, v) for k, v in wf.items() if not k.startswith("__")]

    for node_id, node in node_items:
        ct = node.get("class_type", "")
        inputs = node.get("inputs", {})

        if ct == "CLIPTextEncode":
            if inputs.get("text") == "POSITIVE_PROMPT" and positive_prompt:
                inputs["text"] = positive_prompt
            elif inputs.get("text") == "NEGATIVE_PROMPT" and negative_prompt:
                inputs["text"] = negative_prompt

        elif ct in ("KSampler", "KSamplerSelect"):
            inputs["seed"] = seed

        elif ct == "KSamplerAdvanced":
            if inputs.get("add_noise") == "enable":
                inputs["noise_seed"] = seed

        elif ct in ("EmptyLatentImage", "EmptySD3LatentImage",
                    "EmptyHunyuanLatentVideo", "EmptyLTXVLatentVideo"):
            if width:
                inputs["width"] = width
            if height:
                inputs["height"] = height

        elif ct == "LoadImage":
            img = source_image
            if img:
                inputs["image"] = img

        elif ct in ("TextEncodeQwenImageEditPlus", "TextEncodeQwenImage"):
            if instruction:
                inputs["prompt"] = instruction
            elif positive_prompt:
                inputs["prompt"] = positive_prompt

        elif ct == "CLIPTextEncodeFlux":
            if positive_prompt:
                inputs["clip_l"] = positive_prompt
                inputs["t5xxl"] = positive_prompt

        elif ct == "WanTextEncode":
            if positive_prompt:
                inputs["positive"] = positive_prompt
            if negative_prompt:
                inputs["negative"] = negative_prompt

        elif ct == "RandomNoise":
            inputs["noise_seed"] = seed

    return wf


# ── 从 DB 获取所有已启用工作流 ────────────────────────
def get_enabled_workflows() -> list[WorkflowTestCase]:
    import pymysql
    conn = pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        database=DB_NAME, charset="utf8mb4",
    )
    cur = conn.cursor()
    cur.execute("""
        SELECT id, name, display_name, category, style_tag, test_time, is_enabled
        FROM workflow_template
        WHERE is_enabled = 1
        ORDER BY category, id
    """)
    rows = cur.fetchall()
    conn.close()

    cases = []
    for r in rows:
        cat = r[3]
        needs_image = cat in ("edit", "face", "upscale", "i2v")
        needs_audio = cat == "audio"
        cases.append(WorkflowTestCase(
            db_id=r[0],
            db_name=r[1],
            display_name=r[2],
            category=cat,
            style_tag=r[4],
            needs_image=needs_image,
            needs_audio=needs_audio,
        ))
    return cases


# ── 更新 DB test_time ────────────────────────
def update_test_time(db_id: int, test_time: float):
    import pymysql
    conn = pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        database=DB_NAME, charset="utf8mb4",
    )
    cur = conn.cursor()
    cur.execute(
        "UPDATE workflow_template SET test_time = %s WHERE id = %s",
        (round(test_time, 1), db_id),
    )
    conn.commit()
    conn.close()


# ── 获取测试素材 ────────────────────────
TEST_ASSETS_DIR = BACKEND_DIR.parent / "test-assets"


def find_test_image() -> Optional[Path]:
    """优先使用 test-assets/girl.png（有明确人脸，适合 face 类工作流）"""
    girl = TEST_ASSETS_DIR / "girl.png"
    if girl.exists():
        return girl
    pngs = sorted(AGENT_OUTPUTS_DIR.glob("*.png"))
    if pngs:
        return pngs[0]
    jpgs = sorted(UPLOADS_DIR.glob("*.jpg"))
    if jpgs:
        return jpgs[0]
    return None


def create_masked_test_image(src_path: Path) -> Path:
    """为 inpainting 工作流创建带 alpha 通道的测试图片（中心区域为透明 mask）"""
    import struct, zlib
    # 读取源图尺寸（简单解析 PNG IHDR）
    with open(src_path, "rb") as f:
        data = f.read()
    # PNG IHDR starts at byte 16
    w = int.from_bytes(data[16:20], "big")
    h = int.from_bytes(data[20:24], "big")
    # 创建简单 RGBA PNG：复用原图但加上中心 1/3 区域的 alpha mask
    # 使用简单方法：把原图转 RGBA，中心 1/3 区域 alpha=0
    output = Path("/tmp/test_masked.png")
    try:
        # 尝试用 Pillow
        from PIL import Image
        img = Image.open(src_path).convert("RGBA")
        pixels = img.load()
        cx, cy = img.width // 2, img.height // 2
        rw, rh = img.width // 6, img.height // 6
        for y in range(cy - rh, cy + rh):
            for x in range(cx - rw, cx + rw):
                r, g, b, a = pixels[x, y]
                pixels[x, y] = (r, g, b, 0)
        img.save(output)
    except ImportError:
        # 无 Pillow 则用 raw 方法创建纯色 RGBA PNG
        import io
        size = min(w, h, 512)
        raw_rows = []
        cx, cy = size // 2, size // 2
        rw, rh = size // 6, size // 6
        for y in range(size):
            row = b'\x00'  # filter byte
            for x in range(size):
                if cx - rw <= x < cx + rw and cy - rh <= y < cy + rh:
                    row += b'\xcc\xcc\xcc\x00'  # masked region: transparent
                else:
                    row += b'\xaa\xaa\xaa\xff'  # normal region: opaque grey
            raw_rows.append(row)
        raw = b''.join(raw_rows)
        # Build minimal PNG
        def _chunk(ctype, cdata):
            c = ctype + cdata
            return struct.pack(">I", len(cdata)) + c + struct.pack(">I", zlib.crc32(c) & 0xffffffff)
        ihdr = struct.pack(">IIBBBBB", size, size, 8, 6, 0, 0, 0)
        png = b'\x89PNG\r\n\x1a\n'
        png += _chunk(b'IHDR', ihdr)
        png += _chunk(b'IDAT', zlib.compress(raw))
        png += _chunk(b'IEND', b'')
        output.write_bytes(png)
    return output


def find_test_audio() -> Optional[Path]:
    wavs = sorted(UPLOADS_DIR.glob("*.wav"))
    if wavs:
        return wavs[0]
    return None


# ── 单个 workflow 测试 ────────────────────────
async def test_single_workflow(
    client: SimpleComfyUIClient,
    tc: WorkflowTestCase,
    test_image_name: Optional[str],
) -> WorkflowTestCase:
    """测试单个 workflow，返回更新后的 TestCase"""
    start = time.time()

    try:
        # 1. 加载 workflow JSON
        wf = load_workflow_json(tc.db_name)
        tc.workflow_json = wf

        # 2. 注入参数
        prompt_map = {
            "t2i": "a beautiful ancient Chinese woman in flowing white hanfu, cherry blossoms background, masterpiece, best quality",
            "edit": "change the background to a sunset beach, keep the main subject",
            "face": "a beautiful girl with long black hair, ancient Chinese style, masterpiece",
            "i2v": "gentle camera zoom in, cinematic movement, smooth animation",
            "t2v": "a beautiful waterfall in a lush forest, cinematic, smooth camera movement",
            "upscale": "enhance details",
            "audio": "separate vocals and instruments",
        }
        prompt = prompt_map.get(tc.category, "a beautiful scene, masterpiece")

        params = {
            "positive_prompt": prompt,
            "seed": 42,
            "width": 512,
            "height": 512,
        }

        if tc.needs_image and test_image_name:
            params["source_image"] = test_image_name
        elif tc.needs_image and not test_image_name:
            tc.status = "skip"
            tc.error = "无测试图片"
            tc.duration_s = round(time.time() - start, 2)
            return tc

        # inpainting 工作流使用带 mask 的图片
        if tc.needs_image and "局部清除" in tc.display_name:
            masked_name = getattr(test_single_workflow, '_masked_img', None)
            if masked_name:
                params["source_image"] = masked_name

        if tc.category == "edit":
            params["instruction"] = prompt

        wf = inject_test_params(wf, **params)

        # 3. 执行
        output_type, output_bytes = await client.run_workflow(wf)
        tc.status = "pass"
        tc.output_type = output_type
        tc.output_size = len(output_bytes)

    except FileNotFoundError as e:
        tc.status = "skip"
        tc.error = f"JSON未找到: {e}"
    except TimeoutError as e:
        tc.status = "fail"
        tc.error = str(e)
    except Exception as e:
        tc.status = "fail"
        tc.error = str(e)[:200]

    tc.duration_s = round(time.time() - start, 2)
    return tc


# ── 生成报告 ────────────────────────
def generate_report(cases: list[WorkflowTestCase], gpu_info: str, total_time: float) -> str:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    passed = [c for c in cases if c.status == "pass"]
    failed = [c for c in cases if c.status == "fail"]
    skipped = [c for c in cases if c.status == "skip"]

    lines = [
        "# ComfyUI 远程 Workflow 可用性测试报告",
        "",
        f"> 测试时间: {now}",
        f"> ComfyUI: {COMFYUI_URL}",
        f"> GPU: {gpu_info}",
        f"> 总耗时: {total_time:.1f}s",
        "",
        "## 总览",
        "",
        "| 指标 | 数值 |",
        "|------|------|",
        f"| 总数 | {len(cases)} |",
        f"| ✅ 通过 | {len(passed)} |",
        f"| ❌ 失败 | {len(failed)} |",
        f"| ⏭ 跳过 | {len(skipped)} |",
        f"| 通过率 | {len(passed)/len(cases)*100:.1f}% |",
        "",
        "## 按分类统计",
        "",
        "| 分类 | 总数 | ✅ | ❌ | ⏭ | 平均耗时 |",
        "|------|------|-----|-----|-----|---------|",
    ]

    from collections import Counter
    cats = sorted(set(c.category for c in cases))
    for cat in cats:
        cat_cases = [c for c in cases if c.category == cat]
        cat_pass = [c for c in cat_cases if c.status == "pass"]
        cat_fail = [c for c in cat_cases if c.status == "fail"]
        cat_skip = [c for c in cat_cases if c.status == "skip"]
        avg_time = sum(c.duration_s for c in cat_pass) / len(cat_pass) if cat_pass else 0
        lines.append(f"| {cat} | {len(cat_cases)} | {len(cat_pass)} | {len(cat_fail)} | {len(cat_skip)} | {avg_time:.1f}s |")

    lines += ["", "## 详细结果", ""]
    lines.append("| # | 分类 | 工作流名称 | 风格 | 状态 | 耗时 | 输出 | 错误 |")
    lines.append("|---|------|----------|------|------|------|------|------|")

    for i, c in enumerate(cases, 1):
        status_icon = {"pass": "✅", "fail": "❌", "skip": "⏭", "pending": "⏳"}.get(c.status, "?")
        output = f"{c.output_type} ({c.output_size//1024}KB)" if c.output_size else "-"
        err = c.error[:60] if c.error else "-"
        lines.append(
            f"| {c.db_id} | {c.category} | {c.display_name[:30]} | {c.style_tag or '-'} | "
            f"{status_icon} | {c.duration_s}s | {output} | {err} |"
        )

    # 失败详情
    if failed:
        lines += ["", "## ❌ 失败详情", ""]
        for c in failed:
            lines.append(f"### {c.display_name} (id={c.db_id}, {c.category})")
            lines.append(f"- **风格**: {c.style_tag}")
            lines.append(f"- **耗时**: {c.duration_s}s")
            lines.append(f"- **错误**: `{c.error}`")
            lines.append("")

    # 跳过详情
    if skipped:
        lines += ["", "## ⏭ 跳过详情", ""]
        for c in skipped:
            lines.append(f"- **{c.display_name}** (id={c.db_id}): {c.error}")

    # DB test_time 更新记录
    lines += ["", "## DB test_time 更新记录", ""]
    lines.append("| ID | 工作流 | 耗时 | 已更新 |")
    lines.append("|------|------|------|------|")
    for c in passed:
        lines.append(f"| {c.db_id} | {c.display_name[:30]} | {c.duration_s}s | ✅ |")

    return "\n".join(lines)


# ── 主流程 ────────────────────────
async def main():
    # 清空日志
    with open(LOG_PATH, "w") as f:
        f.write("")

    log("=" * 60)
    log("ComfyUI 远程 Workflow 可用性测试")
    log("=" * 60)

    client = SimpleComfyUIClient(COMFYUI_URL)

    # 1. 健康检查
    log("[1/5] ComfyUI 健康检查...")
    healthy, gpu_info = await client.health_check()
    if not healthy:
        log(f"❌ ComfyUI 不可达: {gpu_info}")
        log("请确认 AutoDL 实例已开机！")
        sys.exit(1)
    log(f"✅ ComfyUI 在线 | GPU: {gpu_info}")

    # 2. 获取工作流列表
    log("[2/5] 获取已启用工作流...")
    cases = get_enabled_workflows()
    log(f"  共 {len(cases)} 个已启用工作流")

    cats_summary = {}
    for c in cases:
        cats_summary[c.category] = cats_summary.get(c.category, 0) + 1
    for cat, cnt in sorted(cats_summary.items()):
        log(f"  - {cat}: {cnt}")

    # 3. 准备测试素材
    log("[3/5] 准备测试素材...")
    test_image_name = None
    test_image_path = find_test_image()
    if test_image_path:
        log(f"  上传测试图片: {test_image_path.name} ({test_image_path.stat().st_size // 1024}KB)")
        try:
            image_bytes = test_image_path.read_bytes()
            test_image_name = await client.upload_image(image_bytes, test_image_path.name)
            log(f"  ✅ 已上传 → ComfyUI: {test_image_name}")
        except Exception as e:
            log(f"  ❌ 上传失败: {e}")
    else:
        log("  ⚠️ 未找到测试图片，需要图片的工作流将被跳过")

    # 上传带 mask 的 inpainting 测试图片
    if test_image_path:
        try:
            masked_path = create_masked_test_image(test_image_path)
            masked_bytes = masked_path.read_bytes()
            masked_name = await client.upload_image(masked_bytes, "test_masked.png")
            test_single_workflow._masked_img = masked_name
            log(f"  ✅ Inpainting mask 图片已上传 → {masked_name}")
        except Exception as e:
            log(f"  ⚠️ Mask 图片创建失败: {e}")
            test_single_workflow._masked_img = None

    test_audio_path = find_test_audio()
    if test_audio_path:
        log(f"  测试音频可用: {test_audio_path.name}")
    else:
        log("  ⚠️ 未找到测试音频")

    # 4. 逐个测试（支持 --only-failed 跳过已通过的）
    skip_ids: set[int] = set()
    if "--only-failed" in sys.argv and Path(REPORT_PATH).exists():
        import re
        report_text = Path(REPORT_PATH).read_text(encoding="utf-8")
        for m in re.finditer(r"\|\s*(\d+)\s*\|.*?\|\s*✅\s*\|", report_text):
            skip_ids.add(int(m.group(1)))
        log(f"  --only-failed: 跳过上轮已通过的 {len(skip_ids)} 个工作流")

    # 排序：轻量模型优先，nunchaku/heavy 放最后（减少 OOM 导致后续全挂）
    def _sort_key(c):
        w = 0
        if c.style_tag and "nunchaku" in c.style_tag:
            w = 2
        elif c.category in ("i2v", "t2v"):
            w = 1
        return (w, c.category, c.db_id)
    cases.sort(key=_sort_key)

    log(f"[4/5] 开始测试 {len(cases)} 个工作流...")
    total_start = time.time()
    consecutive_failures = 0

    for i, tc in enumerate(cases, 1):
        if tc.db_id in skip_ids:
            tc.status = "pass"
            tc.duration_s = 0
            tc.error = "(上轮已通过，本轮跳过)"
            log(f"  [{i}/{len(cases)}] {tc.display_name} → ⏩ 上轮已通过")
            continue

        # 服务崩溃检测：连续 3 个失败且耗时 < 1s → 健康检查 + 等待恢复
        if consecutive_failures >= 3:
            log(f"    ⚠️ 连续 {consecutive_failures} 次快速失败，检查服务...")
            for retry in range(12):  # 最多等 120s
                ok, _ = await client.health_check()
                if ok:
                    log(f"    ✅ 服务已恢复 (等待 {retry * 10}s)")
                    consecutive_failures = 0
                    break
                await asyncio.sleep(10)
            else:
                log(f"    ❌ 服务 120s 未恢复，剩余工作流标记为 skip")
                for tc2 in cases[i - 1:]:
                    if tc2.db_id not in skip_ids and tc2.status == "pending":
                        tc2.status = "skip"
                        tc2.error = "服务崩溃未恢复"
                break

        log(f"  [{i}/{len(cases)}] {tc.display_name} ({tc.category}/{tc.style_tag})...")
        tc = await test_single_workflow(client, tc, test_image_name)

        if tc.status == "fail" and tc.duration_s < 1.0:
            consecutive_failures += 1
        else:
            consecutive_failures = 0

        status_icon = {"pass": "✅", "fail": "❌", "skip": "⏭"}.get(tc.status, "?")
        extra = ""
        if tc.status == "pass":
            extra = f" → {tc.output_type} ({tc.output_size // 1024}KB)"
            # 更新 DB test_time
            try:
                update_test_time(tc.db_id, tc.duration_s)
                extra += " [DB已更新]"
            except Exception as e:
                extra += f" [DB更新失败: {e}]"
        elif tc.error:
            extra = f" | {tc.error[:80]}"

        log(f"    {status_icon} {tc.duration_s}s{extra}")

    total_time = round(time.time() - total_start, 1)

    # 5. 生成报告
    log(f"[5/5] 生成报告...")
    report = generate_report(cases, gpu_info, total_time)
    Path(REPORT_PATH).write_text(report, encoding="utf-8")
    log(f"  报告: {REPORT_PATH}")

    passed = sum(1 for c in cases if c.status == "pass")
    failed = sum(1 for c in cases if c.status == "fail")
    skipped = sum(1 for c in cases if c.status == "skip")

    # 6. 生成 Excel
    log("[6/6] 生成 Excel 报表...")
    generate_excel(cases, gpu_info, total_time)
    log(f"  Excel: {EXCEL_PATH}")

    log("")
    log("=" * 60)
    log(f"测试完成: ✅{passed} ❌{failed} ⏭{skipped} / {len(cases)}")
    log(f"总耗时: {total_time}s")
    log(f"报告: {REPORT_PATH}")
    log(f"Excel: {EXCEL_PATH}")
    log("=" * 60)


def generate_excel(cases: list[WorkflowTestCase], gpu_info: str, total_time: float):
    """生成 Excel 报表（含汇总页+详情页+失败分析页）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── 样式 ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, row, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def auto_width(ws, min_width=8, max_width=50):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    # 中文字符算2个宽度
                    val = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in val)
                    max_len = max(max_len, length)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

    # ════════════════════ Sheet 1: 汇总 ════════════════════
    ws1 = wb.active
    ws1.title = "测试汇总"
    ws1.cell(row=1, column=1, value="ComfyUI 远程 Workflow 可用性测试报告").font = title_font
    ws1.cell(row=2, column=1, value=f"测试时间: {now}")
    ws1.cell(row=3, column=1, value=f"ComfyUI: {COMFYUI_URL}")
    ws1.cell(row=4, column=1, value=f"GPU: {gpu_info}")
    ws1.cell(row=5, column=1, value=f"总耗时: {total_time}s")

    passed = [c for c in cases if c.status == "pass"]
    failed = [c for c in cases if c.status == "fail"]
    skipped = [c for c in cases if c.status == "skip"]

    # 总览表格
    r = 7
    for label, value in [
        ("总工作流数", len(cases)),
        ("✅ 通过", len(passed)),
        ("❌ 失败", len(failed)),
        ("⏭ 跳过", len(skipped)),
        ("通过率", f"{len(passed)/len(cases)*100:.1f}%"),
    ]:
        ws1.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=r, column=2, value=value)
        r += 1

    # 按分类统计
    r += 1
    ws1.cell(row=r, column=1, value="按分类统计").font = Font(bold=True, size=12)
    r += 1
    cat_headers = ["分类", "总数", "通过", "失败", "跳过", "通过率", "平均耗时(s)"]
    for ci, h in enumerate(cat_headers, 1):
        ws1.cell(row=r, column=ci, value=h)
    style_header(ws1, r, len(cat_headers))
    r += 1

    cats = sorted(set(c.category for c in cases))
    for cat in cats:
        cc = [c for c in cases if c.category == cat]
        cp = [c for c in cc if c.status == "pass"]
        cf = [c for c in cc if c.status == "fail"]
        cs = [c for c in cc if c.status == "skip"]
        avg_t = round(sum(c.duration_s for c in cp) / len(cp), 1) if cp else 0
        rate = f"{len(cp)/len(cc)*100:.0f}%" if cc else "0%"
        row_data = [cat, len(cc), len(cp), len(cf), len(cs), rate, avg_t]
        for ci, v in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=ci, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        r += 1

    # 失败原因分类
    r += 1
    ws1.cell(row=r, column=1, value="失败原因分类").font = Font(bold=True, size=12)
    r += 1
    error_cats = {}
    for c in failed:
        if "MarkdownNote" in c.error:
            key = "MarkdownNote 节点缺失"
        elif "Note' not found" in c.error:
            key = "Note 节点缺失"
        elif "scheduler" in c.error:
            key = "参数校验失败"
        elif "超时" in c.error or "Timeout" in c.error:
            key = "执行超时"
        elif "执行失败" in c.error:
            key = "执行失败"
        else:
            key = "其他错误"
        error_cats[key] = error_cats.get(key, 0) + 1

    err_headers = ["错误类型", "数量", "影响占比"]
    for ci, h in enumerate(err_headers, 1):
        ws1.cell(row=r, column=ci, value=h)
    style_header(ws1, r, len(err_headers))
    r += 1
    for ek, ev in sorted(error_cats.items(), key=lambda x: -x[1]):
        ws1.cell(row=r, column=1, value=ek).border = thin_border
        ws1.cell(row=r, column=2, value=ev).border = thin_border
        ws1.cell(row=r, column=3, value=f"{ev/len(cases)*100:.1f}%").border = thin_border
        r += 1

    auto_width(ws1)

    # ════════════════════ Sheet 2: 详细结果 ════════════════════
    ws2 = wb.create_sheet("详细结果")
    detail_headers = [
        "DB_ID", "工作流名称", "显示名称", "分类", "风格",
        "状态", "耗时(s)", "输出类型", "输出大小(KB)",
        "需要图片", "需要音频", "测试素材", "错误原因",
    ]
    for ci, h in enumerate(detail_headers, 1):
        ws2.cell(row=1, column=ci, value=h)
    style_header(ws2, 1, len(detail_headers))

    for ri, c in enumerate(cases, 2):
        status_map = {"pass": "✅通过", "fail": "❌失败", "skip": "⏭跳过"}
        material = ""
        if c.needs_image:
            material = "图片(03f0159d9099.png)"
        if c.needs_audio:
            material = "音频(wav)"
        if not c.needs_image and not c.needs_audio:
            material = "仅文本提示词"

        row_data = [
            c.db_id,
            c.db_name[:60],
            c.display_name,
            c.category,
            c.style_tag or "-",
            status_map.get(c.status, c.status),
            c.duration_s,
            c.output_type or "-",
            c.output_size // 1024 if c.output_size else 0,
            "是" if c.needs_image else "否",
            "是" if c.needs_audio else "否",
            material,
            c.error or "-",
        ]
        for ci, v in enumerate(row_data, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            # 状态着色
            if ci == 6:
                if c.status == "pass":
                    cell.fill = pass_fill
                elif c.status == "fail":
                    cell.fill = fail_fill
                elif c.status == "skip":
                    cell.fill = skip_fill
            cell.alignment = Alignment(vertical="center")

    auto_width(ws2)
    # 冻结首行
    ws2.freeze_panes = "A2"

    # ════════════════════ Sheet 3: 失败分析 ════════════════════
    ws3 = wb.create_sheet("失败分析")
    fail_headers = ["DB_ID", "工作流", "分类", "风格", "耗时(s)", "错误原因", "错误分类", "修复建议"]
    for ci, h in enumerate(fail_headers, 1):
        ws3.cell(row=1, column=ci, value=h)
    style_header(ws3, 1, len(fail_headers))

    for ri, c in enumerate(failed, 2):
        if "MarkdownNote" in c.error:
            err_cat = "节点缺失"
            fix = "在 ComfyUI 安装 MarkdownNote 自定义节点，或从工作流 JSON 中移除该节点"
        elif "Note' not found" in c.error:
            err_cat = "节点缺失"
            fix = "在 ComfyUI 安装 Note 自定义节点"
        elif "scheduler" in c.error:
            err_cat = "参数校验"
            fix = "检查 KSampler 的 scheduler 参数值是否在当前 ComfyUI 版本支持列表中"
        elif "超时" in c.error:
            err_cat = "超时"
            fix = "增大 COMFYUI_TIMEOUT 或检查模型是否已加载"
        elif "执行失败" in c.error:
            err_cat = "执行错误"
            fix = "查看 ComfyUI 服务端日志获取详细错误"
        else:
            err_cat = "其他"
            fix = "需人工排查"

        row_data = [
            c.db_id, c.display_name, c.category, c.style_tag or "-",
            c.duration_s, c.error[:100], err_cat, fix,
        ]
        for ci, v in enumerate(row_data, 1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            cell.fill = fail_fill if ci == 6 else PatternFill()

    auto_width(ws3)
    ws3.freeze_panes = "A2"

    # ════════════════════ Sheet 4: 跳过分析 ════════════════════
    if skipped:
        ws4 = wb.create_sheet("跳过分析")
        skip_headers = ["DB_ID", "工作流", "分类", "风格", "跳过原因", "修复建议"]
        for ci, h in enumerate(skip_headers, 1):
            ws4.cell(row=1, column=ci, value=h)
        style_header(ws4, 1, len(skip_headers))

        for ri, c in enumerate(skipped, 2):
            if "JSON未找到" in c.error:
                fix = "检查 workflow JSON 文件名是否匹配 DB name，或修复 load_by_name 路径解析"
            elif "无测试图片" in c.error:
                fix = "准备测试图片后重试"
            else:
                fix = "需人工排查"
            row_data = [
                c.db_id, c.display_name, c.category, c.style_tag or "-",
                c.error, fix,
            ]
            for ci, v in enumerate(row_data, 1):
                cell = ws4.cell(row=ri, column=ci, value=v)
                cell.border = thin_border
                cell.fill = skip_fill if ci == 5 else PatternFill()

        auto_width(ws4)
        ws4.freeze_panes = "A2"

    wb.save(EXCEL_PATH)
    log(f"  Excel 已保存: {EXCEL_PATH}")


if __name__ == "__main__":
    asyncio.run(main())
